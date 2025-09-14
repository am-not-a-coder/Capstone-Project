import PyPDF2
import docx
import openpyxl


def extract_pdf(file_path):
    text = ""
    with open(file_path, "rb") as f:
        reader = PyPDF2.PdfReader(f)
        for page in reader.pages:
            text += page.extract_text() + "\n"
    return text

def extract_docs(file_path):
    doc = docx.Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs])

def extract_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only = True):
            text += " ".join([str(cell) for cell in row if cell]) + "\n"
    return text
