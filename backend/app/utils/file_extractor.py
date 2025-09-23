import PyPDF2
import docx
import openpyxl
import pytesseract
from pdf2image import convert_from_path
from PIL import Image


def extract_pdf(file_path):
    text = ""
    images = convert_from_path(
        file_path,
        dpi=300
    )
    # Extract text from image inside the pdf
    for i, img in enumerate(images):
        page_text = pytesseract.image_to_string(img, lang="eng")
        text += page_text

    return text

def extract_image(image_path):
    text = pytesseract.image_to_string(Image.open(image_path))
    return text



def extract_docs(file_path):
    doc = docx.Document(file_path)
    return "\n".join([para.text for para in doc.paragraphs])

def extract_excel(file_path):
    workbook = openpyxl.load_workbook(file_path)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only = True):
            text += " ".join([str(cell) for cell in row if cell]) + "\n"
    return text
